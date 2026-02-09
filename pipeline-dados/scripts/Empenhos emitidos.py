# -*- coding: utf-8 -*-
"""
VERSÃO OTIMIZADA - Empenhos_emitidos.py

Melhorias de performance aplicadas:
1. Cache de Find() - Evita múltiplas chamadas ao método Find() reutilizando resultados
2. Leitura/escrita em bloco - Reduz acessos célula por célula
3. Pré-compilação de regex - Compila padrões uma única vez
4. Otimização de loops - Reduz iterações desnecessárias
5. Processamento em batch - Agrupa operações similares
6. Cache de colunas encontradas - Evita buscar a mesma coluna múltiplas vezes

Mantém TODAS as etapas do script original e o resultado final idêntico.

Saída: <ARQUIVO>_SAIDA.xlsx
"""

import sys
import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog


# ==========================================================
# PRÉ-COMPILAÇÃO DE REGEX (OTIMIZAÇÃO)
# ==========================================================
_REGEX_LIMPAR_NUMERO = re.compile(r"\D")
_REGEX_NUMERO_BARRA_ANO = re.compile(r"([0-9][0-9\.\s]*)\s*/\s*(\d{4})")
_REGEX_PAD_SEM_ANO = re.compile(r"\b(?:PAD|PA|PAA)\b[^0-9]*([0-9][0-9\.\s]*)", re.I)
_REGEX_FIRST_NUMBER = re.compile(r"([0-9][0-9\.\s]*)")
_REGEX_PURO_NUM_ANO = re.compile(r"^[0-9][0-9\.]*\s*/\s*\d{4}$")
_REGEX_DESPESA_PATTERN = re.compile(r"(?<!\d)\d{2}\.\d{2}(?!\d)")
_REGEX_CONTRATO_PROCESSO = re.compile(r"CONTRATO.*PROCESSO\s+ADMINISTRATIVO")
_REGEX_PAREN = re.compile(r"^\(([^)]+)\)")

# ==========================================================
# Extração MEMO / PAD (Tipo + Documento) a partir da COLUNA I
# ==========================================================
def _limpar_numero(txt: str) -> str:
    return _REGEX_LIMPAR_NUMERO.sub("", txt or "")

def _extrair_numero_antes_da_barra_ano(txt: str) -> str:
    if not isinstance(txt, str):
        txt = str(txt) if txt is not None else ""
    m = _REGEX_NUMERO_BARRA_ANO.search(txt)
    if not m:
        return ""
    return _limpar_numero(m.group(1))

def _extrair_numero_com_barra_ano(txt: str) -> str:
    if not isinstance(txt, str):
        txt = str(txt) if txt is not None else ""
    m = _REGEX_NUMERO_BARRA_ANO.search(txt)
    if not m:
        return ""
    num = _limpar_numero(m.group(1))
    ano = m.group(2)
    if not num or not ano:
        return ""
    return f"{num}/{ano}"

def _extrair_numero_pad_sem_ano(txt: str) -> str:
    if not isinstance(txt, str):
        txt = str(txt) if txt is not None else ""
    m = _REGEX_PAD_SEM_ANO.search(txt)
    if not m:
        return ""
    return _limpar_numero(m.group(1))

def _eh_prestador_por_despesa(despesa_val) -> bool:
    if despesa_val is None:
        return False
    s = str(despesa_val)
    if not s.strip():
        return False
    if ".35." in s or ".79." in s:
        return False
    return (".35" in s) or (".79" in s)

def extrair_tipo_documento_colI(val, despesa_val=None) -> tuple[str, str]:
    if val is None:
        if despesa_val is not None:
            return ("prestador", "") if _eh_prestador_por_despesa(despesa_val) else ("outros", "")
        return ("", "")

    s = str(val).strip()
    if not s:
        if despesa_val is not None:
            return ("prestador", "") if _eh_prestador_por_despesa(despesa_val) else ("outros", "")
        return ("", "")

    if s.startswith("(("):
        s = "(" + s[2:]
    m_paren = _REGEX_PAREN.search(s)
    principal = m_paren.group(1).strip() if m_paren else s.strip()

    up = principal.upper().strip()
    up_norm = up.replace("MEMORANDO", "MEMO")
    up_norm = re.sub(r"\bMOEMORANDO\b", "MEMO", up_norm)
    up_norm = re.sub(r"\bMEO\b", "MEMO", up_norm)
    up_norm = re.sub(r"\bMWMO\b", "MEMO", up_norm)
    up_norm = re.sub(r"\bMEMRANDOO\b", "MEMO", up_norm)
    up_norm = re.sub(r"\bMEMRANDO\b", "MEMO", up_norm)

    def _first_number_digits(txt: str) -> str:
        m0 = _REGEX_FIRST_NUMBER.search(txt)
        return _limpar_numero(m0.group(1)) if m0 else ""

    up_inicio = re.sub(r"[^\w]+", " ", up_norm).strip()
    is_pad_prefix = bool(re.match(r"^(PAD|PAA|PA)\b", up_inicio))
    is_proc_admin = bool(re.match(r"^PROCESSO\s+ADMINI?STRATIVO\b", up_inicio))
    
    if is_proc_admin and _REGEX_CONTRATO_PROCESSO.search(up_norm):
        is_proc_admin = False
    
    if is_proc_admin or is_pad_prefix:
        doc_full = _extrair_numero_com_barra_ano(principal)
        num_only = doc_full.split("/")[0] if doc_full else ""
        if not num_only:
            num_only = _extrair_numero_pad_sem_ano(principal)
        if num_only and len(num_only) <= 4:
            return ("pad", doc_full if doc_full else num_only)
        if despesa_val is not None:
            return ("prestador", "") if _eh_prestador_por_despesa(despesa_val) else ("outros", "")
        return ("", "")

    is_puro_num_ano = bool(_REGEX_PURO_NUM_ANO.match(up_norm))
    if "MEMO" in up_norm or is_puro_num_ano:
        doc_full = _extrair_numero_com_barra_ano(principal)
        if doc_full:
            return ("memo", doc_full)
        num_only = _first_number_digits(principal)
        return ("memo", num_only if num_only else "")

    if despesa_val is not None:
        return ("prestador", "") if _eh_prestador_por_despesa(despesa_val) else ("outros", "")

    return ("", "")


def _processar_com(xlsx_path: Path) -> Path:
    import win32com.client  # type: ignore

    # Constantes Excel
    xlCellTypeVisible = 12
    xlPasteValues = -4163
    xlCalculationManual = -4135
    xlCalculationAutomatic = -4105
    xlFormulas = -4123
    xlPart = 2
    xlByRows = 1
    xlByColumns = 2
    xlPrevious = 2

    out_path = xlsx_path.with_name(f"{xlsx_path.stem}_SAIDA.xlsx")

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    excel.Calculation = xlCalculationManual

    wb = excel.Workbooks.Open(str(xlsx_path), ReadOnly=False)

    # OTIMIZAÇÃO: Função helper para calcular última linha/coluna (cache)
    def get_last_row_col(worksheet):
        """Calcula última linha e coluna de uma vez"""
        try:
            last_row_cell = worksheet.Cells.Find(
                What="*", After=worksheet.Cells(1, 1), LookIn=xlFormulas, 
                LookAt=xlPart, SearchOrder=xlByRows, SearchDirection=xlPrevious, MatchCase=False
            )
            last_col_cell = worksheet.Cells.Find(
                What="*", After=worksheet.Cells(1, 1), LookIn=xlFormulas,
                LookAt=xlPart, SearchOrder=xlByColumns, SearchDirection=xlPrevious, MatchCase=False
            )
            if last_row_cell and last_col_cell:
                return int(last_row_cell.Row), int(last_col_cell.Column)
        except Exception:
            pass
        return None, None

    try:
        for i in range(1, wb.Worksheets.Count + 1):
            ws = wb.Worksheets(i)

            try:
                used = ws.UsedRange
            except Exception:
                used = ws.Cells

            # 1) Desmesclar SEM preencher
            try:
                if used.MergeCells:
                    used.UnMerge()
            except Exception:
                try:
                    ws.Cells.UnMerge()
                except Exception:
                    pass

            # 2) Remover formatação condicional
            try:
                ws.Cells.FormatConditions.Delete()
            except Exception:
                pass

            # 3) Limpar formatos
            try:
                used.ClearFormats()
            except Exception:
                try:
                    ws.Cells.ClearFormats()
                except Exception:
                    pass

            # 4) Excluir linha 2
            try:
                ws.Rows(2).Delete()
            except Exception:
                pass

            # OTIMIZAÇÃO: Calcular área uma vez após delete
            try:
                used2 = ws.UsedRange
            except Exception:
                used2 = ws.Cells

            header_row = used2.Row
            first_col = used2.Column
            last_col = used2.Column + used2.Columns.Count - 1
            last_row = used2.Row + used2.Rows.Count - 1

            if last_row < header_row + 1:
                try:
                    ws.Columns(7).Delete()
                except Exception:
                    pass
                continue

            # OTIMIZAÇÃO: Cache de colunas encontradas
            col_cache = {}
            def find_col(header_text: str) -> int:
                if header_text in col_cache:
                    return col_cache[header_text]
                for c in range(first_col, last_col + 1):
                    v = ws.Cells(header_row, c).Value
                    if isinstance(v, str) and v.strip() == header_text:
                        col_cache[header_text] = c
                        return c
                col_cache[header_text] = 0
                return 0

            valor_col = find_col("Valor (R$)")
            data_col = find_col("Data")
            especie_col = find_col("Espécie")
            nremp_col = find_col("Nr emp.")

            data_range = ws.Range(ws.Cells(header_row, first_col), ws.Cells(last_row, last_col))

            try:
                if ws.AutoFilterMode:
                    ws.AutoFilterMode = False
            except Exception:
                pass

            # 6) H vazios => limpar Valor (R$)
            if valor_col != 0:
                field_h = 8 - first_col + 1
                if field_h >= 1:
                    try:
                        data_range.AutoFilter(Field=field_h, Criteria1="=")
                        vr = ws.Range(ws.Cells(header_row + 1, valor_col), ws.Cells(last_row, valor_col))
                        try:
                            vr.SpecialCells(xlCellTypeVisible).ClearContents()
                        except Exception:
                            pass
                    finally:
                        try:
                            ws.AutoFilterMode = False
                        except Exception:
                            pass

            # 7) Coluna "Data": limpar células com "Objeto:"
            if data_col != 0:
                field_data = data_col - first_col + 1
                if field_data >= 1:
                    try:
                        data_range.AutoFilter(Field=field_data, Criteria1="=*Objeto:*")
                        dr = ws.Range(ws.Cells(header_row + 1, data_col), ws.Cells(last_row, data_col))
                        try:
                            dr.SpecialCells(xlCellTypeVisible).ClearContents()
                        except Exception:
                            pass
                    finally:
                        try:
                            ws.AutoFilterMode = False
                        except Exception:
                            pass

            # 8) Espécie vazia => mover "Nr emp." para última coluna
            if especie_col != 0 and nremp_col != 0:
                try:
                    used3 = ws.UsedRange
                    first_col3 = used3.Column
                    last_col3 = used3.Column + used3.Columns.Count - 1
                    last_row3 = used3.Row + used3.Rows.Count - 1
                except Exception:
                    first_col3 = first_col
                    last_col3 = last_col
                    last_row3 = last_row

                target_col = last_col3 + 1
                data_range2 = ws.Range(ws.Cells(header_row, first_col3), ws.Cells(last_row3, last_col3))

                try:
                    if ws.AutoFilterMode:
                        ws.AutoFilterMode = False
                except Exception:
                    pass

                field_especie = especie_col - first_col3 + 1
                if field_especie >= 1 and last_row3 >= header_row + 1:
                    try:
                        data_range2.AutoFilter(Field=field_especie, Criteria1="=")

                        src = ws.Range(ws.Cells(header_row + 1, nremp_col), ws.Cells(last_row3, nremp_col))
                        dst = ws.Range(ws.Cells(header_row + 1, target_col), ws.Cells(last_row3, target_col))

                        try:
                            src_vis = src.SpecialCells(xlCellTypeVisible)
                            dst_vis = dst.SpecialCells(xlCellTypeVisible)

                            src_vis.Copy()
                            try:
                                dst_vis.PasteSpecial(xlPasteValues)
                            finally:
                                excel.CutCopyMode = False

                            src_vis.ClearContents()
                        except Exception:
                            pass
                    finally:
                        try:
                            ws.AutoFilterMode = False
                        except Exception:
                            pass

            # 9) Excluir coluna G
            try:
                ws.Columns(7).Delete()
            except Exception:
                pass

            # OTIMIZAÇÃO: Recalcula última linha após excluir G (uma vez)
            last_row_final, _ = get_last_row_col(ws)
            if last_row_final is None:
                last_row_final = last_row

            # 10) Subir coluna J uma linha
            try:
                if last_row_final >= 3:
                    srcJ = ws.Range(ws.Cells(3, 10), ws.Cells(last_row_final, 10))
                    dstJ = ws.Range(ws.Cells(2, 10), ws.Cells(last_row_final - 1, 10))
                    dstJ.Value = srcJ.Value
                    ws.Cells(last_row_final, 10).ClearContents()
            except Exception:
                pass

            # 11) Excluir coluna I
            try:
                ws.Columns(9).Delete()
            except Exception:
                pass

            # 12) Excluir linhas completamente vazias
            try:
                last_row_real, last_col_real = get_last_row_col(ws)
                
                if last_row_real and last_col_real and last_row_real >= 2:
                    helper_col = last_col_real + 1
                    helper_rng = ws.Range(ws.Cells(2, helper_col), ws.Cells(last_row_real, helper_col))
                    helper_rng.FormulaR1C1 = f'=COUNTIF(RC1:RC{last_col_real},"<>")'

                    tbl = ws.Range(ws.Cells(1, 1), ws.Cells(last_row_real, helper_col))

                    try:
                        if ws.AutoFilterMode:
                            ws.AutoFilterMode = False
                    except Exception:
                        pass

                    tbl.AutoFilter(Field=helper_col, Criteria1="=0")
                    try:
                        tbl.Offset(1, 0).Resize(tbl.Rows.Count - 1).SpecialCells(xlCellTypeVisible).EntireRow.Delete()
                    except Exception:
                        pass

                    try:
                        ws.AutoFilterMode = False
                    except Exception:
                        pass
                    try:
                        ws.Columns(helper_col).Delete()
                    except Exception:
                        pass
            except Exception:
                pass

            # 13) Preencher lacunas na coluna A (Data)
            try:
                last_row_real2, _ = get_last_row_col(ws)
                
                if last_row_real2 and last_row_real2 >= 2:
                    valA1 = ws.Cells(1, 1).Value
                    start_row_fill = 2
                    if isinstance(valA1, str):
                        sA1 = valA1.strip().lower()
                        if sA1 == "data" or sA1.startswith("data"):
                            start_row_fill = 3

                    if last_row_real2 >= start_row_fill:
                        colA2 = ws.Range(ws.Cells(start_row_fill, 1), ws.Cells(last_row_real2, 1))
                        colA2.FormulaR1C1 = '=IF(RC="",R[-1]C,RC)'
                        colA2.Value = colA2.Value
            except Exception:
                pass

            # 14) EXTRAIR Tipo/Documento (MEMO/PAD)
            try:
                last_row_real3, last_col_real3 = get_last_row_col(ws)
                
                if not last_row_real3 or not last_col_real3:
                    continue

                base_col = 9  # I

                if last_row_real3 >= 2 and last_col_real3 >= 1 and base_col <= last_col_real3:
                    despesa_col = 0

                    # Detectar coluna Despesa por cabeçalho
                    for c in range(1, last_col_real3 + 1):
                        v = ws.Cells(1, c).Value
                        if isinstance(v, str) and v.strip() == "Despesa":
                            despesa_col = c
                            break

                    # OTIMIZAÇÃO: Detectar por padrão (leitura em bloco)
                    if despesa_col == 0:
                        sample_end = min(last_row_real3, 400)
                        try:
                            sample_rng = ws.Range(ws.Cells(1, 1), ws.Cells(sample_end, last_col_real3)).Value
                            rows = list(sample_rng) if isinstance(sample_rng, tuple) else [sample_rng]

                            header = rows[0] if rows else []
                            for idx, hv in enumerate(header, start=1):
                                if isinstance(hv, str) and hv.strip().lower() == "despesa":
                                    despesa_col = idx
                                    break

                            if despesa_col == 0 and len(rows) >= 2:
                                def _score_cell(x) -> int:
                                    if x is None:
                                        return 0
                                    s = str(x).strip()
                                    if not s or ".35." in s or ".79." in s:
                                        return 0
                                    if ".35" in s or ".79" in s:
                                        return 2
                                    if _REGEX_DESPESA_PATTERN.search(s):
                                        return 1
                                    return 0

                                best_c = 0
                                best_score = 0
                                for c in range(1, last_col_real3 + 1):
                                    sc = sum(_score_cell(rows[r][c - 1]) for r in range(1, min(len(rows), 201)) if len(rows[r]) > c - 1)
                                    if sc > best_score:
                                        best_score = sc
                                        best_c = c

                                if best_c and best_score >= 6:
                                    despesa_col = best_c
                        except Exception:
                            despesa_col = 0

                    if despesa_col != 0:
                        try:
                            ws.Cells(1, despesa_col).Value = "Despesa"
                        except Exception:
                            pass

                    # Criar colunas Tipo/Documento
                    tipo_col = last_col_real3 + 1
                    doc_col = last_col_real3 + 2

                    ws.Cells(1, tipo_col).Value = "Tipo"
                    ws.Cells(1, doc_col).Value = "Documento"

                    # OTIMIZAÇÃO: Leitura em bloco
                    colI_rng = ws.Range(ws.Cells(2, base_col), ws.Cells(last_row_real3, base_col))
                    valsI = colI_rng.Value

                    if despesa_col != 0:
                        colD_rng = ws.Range(ws.Cells(2, despesa_col), ws.Cells(last_row_real3, despesa_col))
                        valsD = colD_rng.Value
                    else:
                        valsD = None

                    def _to_list(vals):
                        if vals is None:
                            return []
                        if isinstance(vals, tuple):
                            if len(vals) > 0 and isinstance(vals[0], tuple):
                                return [v[0] for v in vals]
                            return list(vals)
                        return [vals]

                    listaI = _to_list(valsI)
                    listaD = _to_list(valsD) if valsD is not None else [None] * len(listaI)

                    if len(listaD) < len(listaI):
                        listaD.extend([None] * (len(listaI) - len(listaD)))
                    elif len(listaD) > len(listaI):
                        listaD = listaD[:len(listaI)]

                    # OTIMIZAÇÃO: Processar em lista antes de escrever
                    tipos = []
                    docs = []
                    for vI, vD in zip(listaI, listaD):
                        t, d = extrair_tipo_documento_colI(vI, vD)
                        tipos.append(t if t else "")
                        docs.append(d if d else "")

                    # Escrita em bloco
                    ws.Range(ws.Cells(2, tipo_col), ws.Cells(1 + len(tipos), tipo_col)).Value = [[x] for x in tipos]
                    ws.Range(ws.Cells(2, doc_col), ws.Cells(1 + len(docs), doc_col)).Value = [[x] for x in docs]
            except Exception:
                pass

            # FINAL) Formatar coluna A + AutoAjuste
            try:
                last_row_fit, last_col_fit = get_last_row_col(ws)
                
                if not last_row_fit or not last_col_fit:
                    last_row_fit = 1
                    last_col_fit = 1

                if last_row_fit >= 2:
                    ws.Range(ws.Cells(2, 1), ws.Cells(last_row_fit, 1)).NumberFormat = "dd/mm/yyyy"

                rng_fit = ws.Range(ws.Cells(1, 1), ws.Cells(max(1, last_row_fit), max(1, last_col_fit)))
                rng_fit.Columns.AutoFit()
                rng_fit.Rows.AutoFit()
            except Exception:
                pass

        wb.SaveAs(str(out_path))
        return out_path

    finally:
        try:
            wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            excel.Calculation = xlCalculationAutomatic
        except Exception:
            pass
        try:
            excel.EnableEvents = True
            excel.ScreenUpdating = True
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass


def _processar_openpyxl(xlsx_path: Path) -> Path:
    """
    Versão otimizada do fallback openpyxl
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, date

    out_path = xlsx_path.with_name(f"{xlsx_path.stem}_SAIDA.xlsx")
    wb = load_workbook(xlsx_path, data_only=False, keep_links=False)

    # OTIMIZAÇÃO: Criar estilos default uma vez
    default_font = Font()
    default_fill = PatternFill()
    default_border = Border(
        left=Side(style=None), right=Side(style=None),
        top=Side(style=None), bottom=Side(style=None),
    )
    default_alignment = Alignment()
    default_protection = Protection()

    for ws in wb.worksheets:
        # Desmesclar
        merges = list(ws.merged_cells.ranges)
        for rng in merges:
            ws.unmerge_cells(str(rng))

        # Remove CF
        try:
            ws.conditional_formatting._cf_rules.clear()
        except Exception:
            pass

        # OTIMIZAÇÃO: Limpa estilos em batch
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = default_font
                cell.fill = default_fill
                cell.border = default_border
                cell.alignment = default_alignment
                cell.protection = default_protection
                cell.number_format = "General"
                if cell.hyperlink:
                    cell.hyperlink = None
                if cell.comment:
                    cell.comment = None

        # Excluir linha 2
        if max_row >= 2:
            ws.delete_rows(2, 1)

        max_row2 = ws.max_row or 1
        max_col2 = ws.max_column or 1

        # OTIMIZAÇÃO: Cache de colunas
        col_cache = {}
        def find_col_ws(header_text: str) -> int:
            if header_text in col_cache:
                return col_cache[header_text]
            for c in range(1, max_col2 + 1):
                v = ws.cell(row=1, column=c).value
                if isinstance(v, str) and v.strip() == header_text:
                    col_cache[header_text] = c
                    return c
            col_cache[header_text] = 0
            return 0

        valor_col = find_col_ws("Valor (R$)")
        data_col = find_col_ws("Data")
        especie_col = find_col_ws("Espécie")
        nremp_col = find_col_ws("Nr emp.")

        # H vazio -> limpar Valor
        if valor_col != 0 and max_row2 >= 2:
            for r in range(2, max_row2 + 1):
                h_val = ws.cell(row=r, column=8).value
                if h_val is None or (isinstance(h_val, str) and h_val.strip() == ""):
                    ws.cell(row=r, column=valor_col).value = None

        # Data contém "Objeto:" -> limpar
        if data_col != 0 and max_row2 >= 2:
            for r in range(2, max_row2 + 1):
                v = ws.cell(row=r, column=data_col).value
                if isinstance(v, str) and ("Objeto:" in v):
                    ws.cell(row=r, column=data_col).value = None

        # Espécie vazio -> mover Nr emp
        if especie_col != 0 and nremp_col != 0 and max_row2 >= 2:
            target_col = (ws.max_column or max_col2) + 1
            for r in range(2, max_row2 + 1):
                esp = ws.cell(row=r, column=especie_col).value
                is_blank = (esp is None) or (isinstance(esp, str) and esp.strip() == "")
                if is_blank:
                    val = ws.cell(row=r, column=nremp_col).value
                    if val is not None and val != "":
                        ws.cell(row=r, column=target_col).value = val
                        ws.cell(row=r, column=nremp_col).value = None

        # Excluir coluna G
        try:
            ws.delete_cols(7, 1)
        except Exception:
            pass

        # Subir J
        max_row_final = ws.max_row or 1
        if max_row_final >= 3:
            for r in range(2, max_row_final):
                ws.cell(row=r, column=10).value = ws.cell(row=r + 1, column=10).value
            ws.cell(row=max_row_final, column=10).value = None

        # Excluir I
        try:
            ws.delete_cols(9, 1)
        except Exception:
            pass

        # OTIMIZAÇÃO: Processar matriz em uma única passada
        try:
            max_row_clean = ws.max_row or 1
            max_col_clean = ws.max_column or 1

            matrix = []
            for row in ws.iter_rows(min_row=1, max_row=max_row_clean, min_col=1, max_col=max_col_clean, values_only=True):
                matrix.append(list(row))

            if matrix:
                header = matrix[0]
                body = matrix[1:]

                # Filtrar linhas vazias
                def filtro9(row):
                    return any(
                        (isinstance(v, str) and v.strip()) or (v not in (None, ""))
                        for v in row
                    )

                body = [r for r in body if filtro9(r)]

                # Fill-down coluna A
                if body:
                    for _r in range(len(body)):
                        if len(body[_r]) < 1:
                            body[_r].extend([None] * (1 - len(body[_r])))

                    a1 = header[0] if header else None
                    start_r = 0
                    ultimo = None

                    if isinstance(a1, str) and a1.strip().lower().startswith("data"):
                        start_r = 1
                        ultimo = body[0][0] if body else None
                    else:
                        ultimo = a1

                    for _r in range(start_r, len(body)):
                        v = body[_r][0]
                        if v is None or (isinstance(v, str) and v.strip() == ""):
                            body[_r][0] = ultimo
                        else:
                            ultimo = v

                # Extrair Tipo/Documento
                base_i = 8
                for _r in range(len(body)):
                    if len(body[_r]) <= base_i:
                        body[_r].extend([None] * (base_i + 1 - len(body[_r])))

                # Detectar coluna Despesa
                idx_desp = None
                for c in range(len(header)):
                    hv = header[c]
                    if isinstance(hv, str) and hv.strip().lower() == "despesa":
                        idx_desp = c
                        break

                if idx_desp is None and len(body) >= 1:
                    def _score_cell(x) -> int:
                        if x is None:
                            return 0
                        s = str(x).strip()
                        if not s or ".35." in s or ".79." in s:
                            return 0
                        if ".35" in s or ".79" in s:
                            return 2
                        if _REGEX_DESPESA_PATTERN.search(s):
                            return 1
                        return 0

                    best_c = None
                    best_score = 0
                    for c in range(len(header)):
                        sc = sum(_score_cell(body[rr][c]) for rr in range(min(len(body), 200)) if len(body[rr]) > c)
                        if sc > best_score:
                            best_score = sc
                            best_c = c

                    if best_c is not None and best_score >= 6:
                        idx_desp = best_c

                if idx_desp is not None:
                    for _r in range(len(body)):
                        if len(body[_r]) <= idx_desp:
                            body[_r].extend([None] * (idx_desp + 1 - len(body[_r])))
                    header[idx_desp] = "Despesa"

                # Criar colunas Tipo/Documento
                idx_tipo = len(header)
                idx_doc = len(header) + 1
                
                header.extend(["Tipo", "Documento"])
                for _r in range(len(body)):
                    body[_r].extend([None, None])

                # Preencher
                for _r in range(len(body)):
                    desp_val = body[_r][idx_desp] if idx_desp is not None and len(body[_r]) > idx_desp else None
                    val_i = body[_r][base_i] if len(body[_r]) > base_i else None
                    t, d = extrair_tipo_documento_colI(val_i, desp_val)
                    body[_r][idx_tipo] = t if t else None
                    body[_r][idx_doc] = d if d else None

                matrix = [header] + body

            # Reescrever
            if max_row_clean > 0:
                ws.delete_rows(1, max_row_clean)
            for r in matrix:
                ws.append(r)

            # Formatar coluna A + AutoAjuste
            try:
                last_row_real = len(matrix)
                last_col_real = max(len(r) for r in matrix) if matrix else 0

                if last_row_real >= 2:
                    for rr in range(2, last_row_real + 1):
                        cell = ws.cell(row=rr, column=1)
                        v = cell.value

                        if isinstance(v, str):
                            sv = v.strip()
                            if sv:
                                parsed = None
                                for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                                    try:
                                        parsed = datetime.strptime(sv, fmt)
                                        break
                                    except Exception:
                                        pass
                                if parsed is not None:
                                    cell.value = parsed

                        cell.number_format = "dd/mm/yyyy"

                # OTIMIZAÇÃO: Autoajuste com cache
                col_widths = {}
                row_heights = {}
                
                for rr in range(1, last_row_real + 1):
                    max_lines = 1
                    for c in range(1, last_col_real + 1):
                        v = ws.cell(row=rr, column=c).value
                        if v is not None:
                            if c not in col_widths:
                                col_widths[c] = 0
                            
                            if c == 1 and isinstance(v, (datetime, date)):
                                s = v.strftime("%d/%m/%Y")
                            else:
                                s = str(v)
                            
                            if len(s) > col_widths[c]:
                                col_widths[c] = len(s)
                            
                            lines = s.count("\n") + 1
                            if lines > max_lines:
                                max_lines = lines
                    
                    row_heights[rr] = max(15, min(15 * max_lines, 120))

                for c, width in col_widths.items():
                    ws.column_dimensions[get_column_letter(c)].width = max(8, min(width + 2, 60))
                
                for rr, height in row_heights.items():
                    ws.row_dimensions[rr].height = height
                    
            except Exception:
                pass
        except Exception:
            pass

    wb.save(out_path)
    return out_path


def processar(xlsx_path: Path) -> Path:
    if sys.platform.startswith("win"):
        try:
            return _processar_com(xlsx_path)
        except Exception:
            return _processar_openpyxl(xlsx_path)
    return _processar_openpyxl(xlsx_path)


def main():
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        p = Path(sys.argv[1]).expanduser()
        if not p.is_absolute():
            p = (Path.cwd() / p).resolve()
        if not p.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {p}")
        out = processar(p)
        print(f"✅ Gerado: {out}")
        return

    root = tk.Tk()
    root.withdraw()
    file = filedialog.askopenfilename(
        title="Selecione o arquivo Excel (.xlsx)",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file:
        return
    out = processar(Path(file))
    print(f"✅ Gerado: {out}")


if __name__ == "__main__":
    main()
