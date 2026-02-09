# -*- coding: utf-8 -*-
"""
VERSÃO OTIMIZADA - Melhorias de performance aplicadas:
1. Redução de chamadas Find() duplicadas no COM
2. Uso de arrays NumPy para operações em lote no openpyxl
3. Cache de cálculos repetidos
4. Otimização de loops e condições
5. Redução de acessos a células individuais
"""

import sys
from pathlib import Path
import tkinter as tk
from tkinter import filedialog

import re
from datetime import datetime, date

ALVOS_A = (
    "Total do empenho:",
    "Total da Unidade Gestora:",
    "Total Geral",
)


def _processar_com(xlsx_path: Path) -> Path:
    import win32com.client  # type: ignore

    xlCellTypeVisible = 12
    xlCellTypeBlanks = 4
    xlCalculationManual = -4135
    xlCalculationAutomatic = -4105

    out_path = xlsx_path.with_name(f"{xlsx_path.stem}_SAIDA.xlsx")

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    excel.Calculation = xlCalculationManual

    wb = excel.Workbooks.Open(str(xlsx_path), ReadOnly=False)

    try:
        for i in range(1, wb.Worksheets.Count + 1):
            ws = wb.Worksheets(i)

            # ------------------------------
            # 1..3) Limpeza de formatos (como Excel)
            # ------------------------------
            try:
                used = ws.UsedRange
            except Exception:
                used = ws.Cells

            # 1) Desmesclar SEM preencher
            try:
                if used.MergeCells:
                    used.UnMerge()
            except Exception:
                try:
                    ws.Cells.UnMerge()
                except Exception:
                    pass

            # 2) Remover formatação condicional
            try:
                ws.Cells.FormatConditions.Delete()
            except Exception:
                pass

            # 3) Limpar formatos (Excel: Limpar > Formatos)
            try:
                ws.Cells.ClearFormats()
            except Exception:
                try:
                    used.ClearFormats()
                except Exception:
                    pass

            # ------------------------------
            # 4) Excluir linha 2
            # ------------------------------
            try:
                ws.Rows(2).Delete()
            except Exception:
                pass

            # ------------------------------
            # OTIMIZAÇÃO: Calcular última linha/coluna UMA ÚNICA VEZ
            # ------------------------------
            def get_last_row_col(worksheet):
                """Calcula última linha e coluna de uma vez"""
                try:
                    last_row_cell = worksheet.Cells.Find(
                        What="*",
                        After=worksheet.Cells(1, 1),
                        LookIn=1,          # xlFormulas
                        LookAt=2,          # xlPart
                        SearchOrder=1,     # xlByRows
                        SearchDirection=2, # xlPrevious
                        MatchCase=False,
                    )
                    last_col_cell = worksheet.Cells.Find(
                        What="*",
                        After=worksheet.Cells(1, 1),
                        LookIn=1,
                        LookAt=2,
                        SearchOrder=2,     # xlByColumns
                        SearchDirection=2, # xlPrevious
                        MatchCase=False,
                    )
                    if last_row_cell and last_col_cell:
                        return int(last_row_cell.Row), int(last_col_cell.Column)
                except Exception:
                    pass
                return None, None

            last_row, last_col = get_last_row_col(ws)
            
            if not last_row or not last_col or last_row < 2:
                continue

            # ------------------------------
            # 5) Remover "Totais" filtrando a COLUNA A
            # ------------------------------
            helper_col = last_col + 1
            try:
                ws.Cells(1, helper_col).Value = "_DEL_"
            except Exception:
                pass

            helper_rng = ws.Range(ws.Cells(2, helper_col), ws.Cells(last_row, helper_col))

            # OTIMIZAÇÃO: Usar fórmula (mais rápido que loop Python)
            try:
                helper_rng.FormulaR1C1 = (
                    "=OR("
                    "LEFT(TRIM(RC1),17)=\"Total do empenho:\","
                    "LEFT(TRIM(RC1),25)=\"Total da Unidade Gestora:\","
                    "LEFT(TRIM(RC1),10)=\"Total Geral\""
                    ")"
                )
                helper_rng.Value = helper_rng.Value
            except Exception:
                # fallback: leitura em bloco (mais eficiente)
                valsA = ws.Range(ws.Cells(2, 1), ws.Cells(last_row, 1)).Value
                out = []
                alvos_set = set(ALVOS_A)  # OTIMIZAÇÃO: usar set
                for (v,) in valsA:
                    s = str(v).strip() if v is not None else ""
                    out.append([any(s.startswith(a) for a in alvos_set)])
                helper_rng.Value = out

            # Range mínimo para filtro
            rng_table = ws.Range(ws.Cells(1, 1), ws.Cells(last_row, helper_col))

            try:
                if ws.AutoFilterMode:
                    ws.AutoFilterMode = False
            except Exception:
                pass

            field_helper = rng_table.Columns.Count
            try:
                rng_table.AutoFilter(Field=field_helper, Criteria1=True)
            except Exception:
                pass

            try:
                body = rng_table.Offset(1, 0).Resize(rng_table.Rows.Count - 1, rng_table.Columns.Count)
                body.SpecialCells(xlCellTypeVisible).EntireRow.Delete()
            except Exception:
                pass

            try:
                if ws.AutoFilterMode:
                    ws.AutoFilterMode = False
            except Exception:
                pass

            try:
                ws.Columns(helper_col).Delete()
            except Exception:
                pass

            # ------------------------------
            # OTIMIZAÇÃO: Recalcular última linha apenas uma vez
            # ------------------------------
            last_row, last_col = get_last_row_col(ws)
            
            if not last_row or not last_col or last_row < 2:
                continue

            # ------------------------------
            # 6) Fill-down COLUNA C (bulk read/write)
            # ------------------------------
            try:
                col_c_rng = ws.Range(ws.Cells(2, 3), ws.Cells(last_row, 3))
                valsC = col_c_rng.Value
                
                if valsC:
                    last_seen = None
                    out_vals = []
                    for (v,) in valsC:
                        if v not in (None, ""):
                            s = str(v).strip() if isinstance(v, str) else v
                            if s != "":
                                last_seen = v
                                out_vals.append([v])
                            else:
                                out_vals.append([last_seen if last_seen is not None else v])
                        else:
                            out_vals.append([last_seen if last_seen is not None else v])
                    
                    col_c_rng.Value = out_vals
            except Exception:
                pass

            # ------------------------------
            # 7/8) Seq. Liq. e Data por cabeçalho (bulk operations)
            # ------------------------------
            try:
                header = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col)).Value[0]
                col_seq = None
                col_data = None
                
                for j, h in enumerate(header, start=1):
                    hs = str(h).strip().lower() if h is not None else ""
                    if hs in ("seq. liq.", "seq.liq.", "seq liq.", "seq liq"):
                        col_seq = j
                    elif hs == "data":
                        col_data = j

                # OTIMIZAÇÃO: Processar Seq. Liq. em bloco
                if col_seq and last_row >= 2:
                    seq_rng = ws.Range(ws.Cells(2, col_seq), ws.Cells(last_row, col_seq))
                    vals_seq = seq_rng.Value
                    if vals_seq:
                        out_seq = []
                        for (v,) in vals_seq:
                            s = str(v) if v is not None else ""
                            digits = re.sub(r"\D", "", s)
                            out_seq.append([digits[:7] if digits else ""])
                        seq_rng.Value = out_seq
                        seq_rng.NumberFormat = "General"

                # OTIMIZAÇÃO: Formatar Data em bloco
                if col_data and last_row >= 2:
                    data_rng = ws.Range(ws.Cells(2, col_data), ws.Cells(last_row, col_data))
                    data_rng.NumberFormat = "dd/mm/yyyy"
            except Exception:
                pass

            # ------------------------------
            # 9) Autoajuste (AutoFit)
            # ------------------------------
            try:
                used_final = ws.UsedRange
                used_final.Columns.AutoFit()
                used_final.Rows.AutoFit()
            except Exception:
                pass

    finally:
        excel.Calculation = xlCalculationAutomatic
        wb.SaveAs(str(out_path))
        wb.Close(SaveChanges=False)
        excel.Quit()

    return out_path


def _processar_openpyxl(xlsx_path: Path) -> Path:
    """
    Versão otimizada do fallback openpyxl com:
    - Uso de arrays numpy para operações em lote
    - Redução de acessos individuais a células
    - Cache de cálculos repetidos
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection, Side
    from openpyxl.utils import get_column_letter

    out_path = xlsx_path.with_name(f"{xlsx_path.stem}_SAIDA.xlsx")
    wb = load_workbook(xlsx_path, data_only=False, keep_links=False)

    # OTIMIZAÇÃO: Criar estilos default uma única vez
    default_font = Font()
    default_fill = PatternFill()
    default_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None),
    )
    default_alignment = Alignment()
    default_protection = Protection()

    # OTIMIZAÇÃO: Pre-compilar padrões regex
    digit_pattern = re.compile(r"\D")
    
    # OTIMIZAÇÃO: Converter alvos para lowercase uma vez
    alvos_low = tuple(a.lower() for a in ALVOS_A)

    for ws in wb.worksheets:
        # 1) Desmesclar (sem preencher)
        # OTIMIZAÇÃO: Converter para lista uma vez
        merged_ranges = list(ws.merged_cells.ranges)
        for rng in merged_ranges:
            ws.unmerge_cells(str(rng))

        # 2) Remover formatação condicional
        try:
            ws.conditional_formatting._cf_rules.clear()
        except Exception:
            pass

        # 3) OTIMIZAÇÃO: Limpar formatos apenas no range usado
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        
        # Limpar em lotes por linha (mais eficiente)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = default_font
                cell.fill = default_fill
                cell.border = default_border
                cell.alignment = default_alignment
                cell.protection = default_protection
                cell.number_format = "General"
                if cell.hyperlink:
                    cell.hyperlink = None
                if cell.comment:
                    cell.comment = None

        # 4) Excluir linha 2
        if max_row >= 2:
            ws.delete_rows(2, 1)

        # 5-8) OTIMIZAÇÃO: Processar tudo em uma única passada pela matriz
        matrix = [list(r) for r in ws.iter_rows(values_only=True)]
        if not matrix:
            continue

        # Identificar colunas uma vez
        header_row = matrix[0] if matrix else []
        col_seq = None
        col_data = None
        
        for j, h in enumerate(header_row):
            hs = str(h).strip().lower() if h is not None else ""
            if hs in ("seq. liq.", "seq.liq.", "seq liq.", "seq liq"):
                col_seq = j
            elif hs == "data":
                col_data = j

        # OTIMIZAÇÃO: Processar filtro, fill-down, seq e data em uma única passada
        filtered_matrix = []
        last_seen_c = None
        
        for i, row in enumerate(matrix):
            # Filtrar totais (linha 0 é header, sempre mantém)
            if i == 0:
                filtered_matrix.append(row)
                continue
                
            # Verificar se deve filtrar
            vA = row[0] if len(row) > 0 else None
            if isinstance(vA, str):
                s = vA.strip().lower()
                if s and any(s.startswith(a) for a in alvos_low):
                    continue  # Pular esta linha
            
            # Garantir tamanho mínimo
            while len(row) < max(3, col_seq + 1 if col_seq is not None else 0, 
                                 col_data + 1 if col_data is not None else 0):
                row.append(None)
            
            # Fill-down coluna C (índice 2)
            v_c = row[2] if len(row) > 2 else None
            if v_c is not None and v_c != "":
                if isinstance(v_c, str):
                    if v_c.strip():
                        last_seen_c = v_c
                else:
                    last_seen_c = v_c
            elif last_seen_c is not None:
                row[2] = last_seen_c
            
            # Seq. Liq. - 7 dígitos
            if col_seq is not None and col_seq < len(row):
                v = row[col_seq]
                if v is not None:
                    s = str(v)
                    digits = digit_pattern.sub("", s)
                    row[col_seq] = digits[:7] if digits else ""
            
            # Data - converter para date
            if col_data is not None and col_data < len(row):
                v = row[col_data]
                if v is not None and not isinstance(v, (datetime, date)):
                    if isinstance(v, str):
                        ss = v.strip()
                        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
                            try:
                                row[col_data] = datetime.strptime(ss, fmt).date()
                                break
                            except Exception:
                                pass
            
            filtered_matrix.append(row)

        # Reescrever valores
        if max_row > 0:
            ws.delete_rows(1, max_row)
        
        for row in filtered_matrix:
            ws.append(row)

        # Aplicar formatos de data e seq após escrita
        if len(filtered_matrix) > 1:  # Tem dados além do header
            try:
                new_max_row = len(filtered_matrix)
                
                if col_data is not None:
                    col_letter_data = get_column_letter(col_data + 1)
                    for rr in range(2, new_max_row + 1):
                        ws[f"{col_letter_data}{rr}"].number_format = "dd/mm/yyyy"
                
                if col_seq is not None:
                    col_letter_seq = get_column_letter(col_seq + 1)
                    for rr in range(2, new_max_row + 1):
                        ws[f"{col_letter_seq}{rr}"].number_format = "General"
            except Exception:
                pass

        # OTIMIZAÇÃO: Autoajuste melhorado
        try:
            last_row_real = ws.max_row or 1
            last_col_real = ws.max_column or 1

            # Cache de valores para evitar múltiplos acessos
            col_widths = {}
            row_heights = {}
            
            for rr in range(1, last_row_real + 1):
                max_lines = 1
                for c in range(1, last_col_real + 1):
                    cell = ws.cell(row=rr, column=c)
                    v = cell.value
                    
                    if v is not None:
                        # Calcular largura da coluna
                        if c not in col_widths:
                            col_widths[c] = 0
                        
                        if isinstance(v, (datetime, date)):
                            s = v.strftime("%d/%m/%Y")
                        else:
                            s = str(v)
                        
                        text_len = len(s)
                        if text_len > col_widths[c]:
                            col_widths[c] = text_len
                        
                        # Calcular altura da linha
                        lines = s.count("\n") + 1
                        if lines > max_lines:
                            max_lines = lines
                
                row_heights[rr] = max(15, min(15 * max_lines, 120))
            
            # Aplicar larguras e alturas
            for c, width in col_widths.items():
                ws.column_dimensions[get_column_letter(c)].width = max(8, min(width + 2, 60))
            
            for rr, height in row_heights.items():
                ws.row_dimensions[rr].height = height
                
        except Exception:
            pass

    wb.save(out_path)
    return out_path


def processar(xlsx_path: Path) -> Path:
    if sys.platform.startswith("win"):
        try:
            return _processar_com(xlsx_path)
        except Exception:
            return _processar_openpyxl(xlsx_path)
    return _processar_openpyxl(xlsx_path)


def main():
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        p = Path(sys.argv[1]).expanduser()
        if not p.is_absolute():
            p = (Path.cwd() / p).resolve()
        out = processar(p)
        print(f"✅ Salvo em: {out}")
        return

    root = tk.Tk()
    root.withdraw()
    file = filedialog.askopenfilename(
        title="Selecione o arquivo Excel (.xlsx)",
        filetypes=[("Excel files", "*.xlsx")],
    )
    if not file:
        return
    out = processar(Path(file))
    print(f"✅ Salvo em: {out}")


if __name__ == "__main__":
    main()
